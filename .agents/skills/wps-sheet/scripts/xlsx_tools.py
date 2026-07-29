#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""WPS 表格助手 · 用 openpyxl 生成 / 读取 .xlsx(WPS 表格与 Excel 通用格式)。

依赖:pip install openpyxl
既可被 Codex 当函数库调用,也可直接运行生成一份示例销售表:
    python xlsx_tools.py 销售表.xlsx
"""
from __future__ import annotations
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def new_workbook(sheet_name="Sheet1"):
    """新建工作簿,返回 (wb, ws)。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def write_table(ws, header, rows, start_row=1):
    """写入表头 + 数据,并给表头加粗+底色+边框。返回数据结束行号。"""
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="111111")
    for c, name in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    r = start_row
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
    return r


def autofit(ws, max_width=40):
    """按内容粗略自适应列宽。"""
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 4, max_width)


def set_formula(ws, cell, formula):
    """写入公式,如 set_formula(ws, 'D2', '=B2*C2')。"""
    ws[cell] = formula


def read_sheet(path, sheet=None):
    """读取表格为二维列表(values_only)。"""
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet] if sheet else wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def build_demo(path="示例-销售表.xlsx"):
    """生成一份带公式与合计的示例销售表,验证依赖与流程。"""
    wb, ws = new_workbook("一季度销售")
    header = ["产品", "单价(元)", "数量", "金额(元)"]
    data = [["键盘", 199, 120], ["鼠标", 89, 240], ["显示器", 1299, 35]]
    end = write_table(ws, header, [r + [None] for r in data])  # 金额列留空,用公式
    for i in range(len(data)):
        row = i + 2  # 数据从第 2 行开始
        set_formula(ws, f"D{row}", f"=B{row}*C{row}")
    # 合计行
    total_row = end + 1
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws[f"C{total_row}"] = f"=SUM(C2:C{end})"
    ws[f"D{total_row}"] = f"=SUM(D2:D{end})"
    ws[f"C{total_row}"].font = Font(bold=True)
    ws[f"D{total_row}"].font = Font(bold=True)
    autofit(ws)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "示例-销售表.xlsx"
    saved = build_demo(out)
    print("已生成 WPS/Excel 表格:", saved)
    print("--- 读回校验(前 2 行)---")
    for row in read_sheet(saved)[:2]:
        print(row)
