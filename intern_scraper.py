#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
互联网大厂 2027届实习岗位信息爬虫（优化版）
每天自动抓取字节、腾讯、阿里、百度、美团、快手、拼多多、小红书 的实习岗位信息

优化点：
1. 浏览器实例复用 —— 一次启动，各公司独立 page，最后统一关闭（大幅提速）
2. 页面加载重试机制 —— 网络波动自动重试
3. 字节跳动滚动加载 —— 滚到底部加载更多岗位
4. 邮件支持 Excel 附件
5. 统一的文本解析辅助函数，减少重复代码

使用方法：
    python intern_scraper.py                          # 输出到控制台
    python intern_scraper.py --output report.md       # 输出 Markdown
    python intern_scraper.py --output report.xlsx     # 输出 Excel
    python intern_scraper.py --output-excel report.xlsx --send-email  # 输出Excel并带附件发邮件
    python intern_scraper.py --company 字节跳动,小红书  # 只抓取指定公司
"""

import sys
import os
import re
import time
import argparse
from datetime import datetime
from dataclasses import dataclass
from typing import Optional

# ── 确保输出编码为 UTF-8 ──
if sys.stdout.encoding != 'utf-8':
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

EDGE_PATH = r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe'

CITIES = ["北京", "上海", "深圳", "杭州", "广州", "成都", "武汉", "南京", "珠海", "厦门", "西安", "重庆"]

# 公司展示顺序
COMPANY_ORDER = ["字节跳动", "腾讯", "阿里巴巴", "百度", "美团", "快手", "拼多多", "小红书"]


# ─── 数据模型 ───────────────────────────────────────────────

@dataclass
class InternPosition:
    company: str          # 公司名称
    title: str            # 岗位名称
    location: str         = "未标注"   # 工作地点
    requirement: str      = "待查看"   # 岗位要求（精简版）
    salary: str           = "未公开"   # 薪资范围
    program: str          = ""         # 所属项目
    link: str             = ""         # 详情链接
    source_note: str      = ""         # 数据来源备注


# ─── 浏览器管理 ─────────────────────────────────────────────

def _launch_browser():
    """启动浏览器（Windows 用 Edge，Linux 用 Playwright Chromium）"""
    from playwright.sync_api import sync_playwright
    p = sync_playwright().start()
    if sys.platform == "win32":
        browser = p.chromium.launch(executable_path=EDGE_PATH, headless=True)
    else:
        browser = p.chromium.launch(headless=True)
    return p, browser


class BrowserSession:
    """浏览器会话上下文管理器：一次启动，多处复用，最后统一关闭"""
    def __init__(self):
        self._p = None
        self.browser = None

    def __enter__(self):
        self._p, self.browser = _launch_browser()
        return self

    def new_page(self):
        return self.browser.new_page()

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            self.browser.close()
        finally:
            self._p.stop()


# ─── 通用工具 ───────────────────────────────────────────────

def _page_text(page) -> list:
    """获取页面所有非空行文本"""
    text = page.inner_text("body")
    return [l.strip() for l in text.split("\n") if l.strip()]


def _extract_cities(line: str) -> str:
    """从一行文本中提取城市列表"""
    found = [c for c in CITIES if c in line]
    return "/".join(found[:3]) if found else ""


def _goto_with_retry(page, url, retries=2, wait_ms=3000, wait_until="domcontentloaded", timeout=20000):
    """带重试的页面加载"""
    for attempt in range(retries + 1):
        try:
            page.goto(url, wait_until=wait_until, timeout=timeout)
            page.wait_for_timeout(wait_ms)
            return True
        except Exception as e:
            if attempt >= retries:
                raise
            print(f"    ⚠ 加载失败({attempt+1}/{retries})，重试中... {str(e)[:50]}")
            time.sleep(2)
    return False


def _scroll_to_bottom(page, step_ms=300, max_steps=20):
    """滚动到底部触发懒加载"""
    for _ in range(max_steps):
        page.mouse.wheel(0, 1500)
        page.wait_for_timeout(step_ms)


# ─── 各公司爬虫（均接收 browser 参数，复用实例）─────────────────

def scrape_xiaohongshu(session: BrowserSession) -> list:
    """小红书 - Ace精英实习生岗位（点击"实习生"筛选后解析）"""
    results = []
    seen = set()
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://job.xiaohongshu.com/campus/position", wait_ms=4000)

        # 点击"实习生"筛选，只保留实习岗位
        try:
            for sel in ['text=实习生', 'text=Ace顶尖实习生', 'text=Ace 顶尖实习生']:
                el = page.query_selector(sel)
                if el:
                    el.click()
                    page.wait_for_timeout(2500)
        except Exception:
            pass

        lines = _page_text(page)
        skip = {"首页", "REDstar 顶尖校招", "Ace 精英实习生", "职位", "校招须知",
                "社会招聘", "了解小红书", "登录", "注册", "筛选", "清空", "立即投递",
                "招聘类型", "应届生", "实习生", "Ace精英实习生", "Ace 顶尖实习生", "Ace顶尖实习生",
                "招聘项目", "REDstar 顶尖人才计划", "2026 春季校园招聘", "马当路练习生",
                "职位类别", "技术类", "产品类", "运营类", "设计类", "销售类", "职能类", "城市",
                "北京市", "上海市", "广州市", "深圳市", "武汉市", "杭州市",
                "南京市", "珠海市", "成都市", "其他",
                "Ace 顶尖实习生计划", "Ace顶尖实习生计划"}

        i = 0
        while i < len(lines):
            line = lines[i]
            if line in skip or line.startswith("全部职位") or line.startswith("共") or "个职位" in line:
                i += 1
                continue

            # 岗位行：以【开头（如 【Ace顶尖实习生】xxx / 【27届实习】xxx）
            if line.startswith("【") and len(line) > 10:
                title = line
                location = ""
                req = ""
                j = i + 1
                while j < min(i + 6, len(lines)):
                    nxt = lines[j]
                    cities = _extract_cities(nxt)
                    if cities:
                        location = cities
                    elif len(nxt) > 15 and any(k in nxt for k in ("研究", "开发", "优化", "系统", "算法", "模型", "工程", "平台", "负责", "职责", "课题")):
                        req = nxt[:100] + ("..." if len(nxt) > 100 else "")
                        break
                    j += 1

                clean_title = re.sub(r'[【】]', '', title)[:25]
                if clean_title not in seen:
                    seen.add(clean_title)
                    results.append(InternPosition(
                        company="小红书", title=title,
                        location=location or "北京/上海/杭州",
                        requirement=req or "详见官网",
                        program="Ace精英实习生计划",
                        link="https://job.xiaohongshu.com/campus/position",
                    ))
            i += 1

        # 兜底：正则匹配
        if len(results) < 3:
            results.clear()
            pattern = r'(【[^】]+】[\u4e00-\u9fffA-Za-z0-9（）()-]{4,60})'
            for m in re.findall(pattern, "\n".join(lines)):
                if m not in seen and len(m) > 6:
                    seen.add(m)
                    results.append(InternPosition(
                        company="小红书", title=m,
                        location="北京/上海/杭州等",
                        link="https://job.xiaohongshu.com/campus/position",
                    ))
    except Exception as e:
        results.append(InternPosition(company="小红书", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


def scrape_tencent(session: BrowserSession) -> list:
    """腾讯 - 实习项目"""
    results = []
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://join.qq.com/", wait_ms=3000)
        lines = _page_text(page)

        skip = ["毕业时间", "工作地点在中国", "首页", "登录", "了解腾讯", "帮助中心", "隐私政策"]
        seen_norm = set()
        for line in lines:
            if any(s in line for s in skip):
                continue
            if "实习" in line and "招聘" in line:
                # 归一化去重：去掉"腾讯"、年份、"实习"/"招聘"后比较
                norm = re.sub(r'腾讯|2026|2027|实习|招聘', '', line).strip()
                if norm and norm not in seen_norm:
                    seen_norm.add(norm)
                    results.append(InternPosition(
                        company="腾讯", title=line,
                        location="深圳/北京/上海/广州/成都",
                        requirement="详见腾讯校招官网",
                        link="https://join.qq.com/",
                        source_note="招聘项目信息（具体岗位需在官网查看）",
                    ))

        for kw in ["技术大咖", "Pre留学生"]:
            if any(kw in r.title for r in results):
                continue
            for line in lines:
                if kw in line and ("实习" in line or "培训" in line or "计划" in line):
                    results.append(InternPosition(
                        company="腾讯", title=f"{kw}相关实习项目",
                        location="深圳/北京/上海/广州/成都",
                        link="https://join.qq.com/",
                    ))
                    break

        if not results:
            results.append(InternPosition(
                company="腾讯", title="腾讯2026实习招聘（面向2027届）",
                location="深圳/北京/上海/广州/成都",
                link="https://join.qq.com/",
                source_note="招聘项目已开放，具体岗位请登录官网查看",
            ))
    except Exception as e:
        results.append(InternPosition(company="腾讯", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


def scrape_pinduoduo(session: BrowserSession) -> list:
    """拼多多 - 2027届校招/实习"""
    results = []
    seen = set()
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://careers.pinduoduo.com/campus", wait_ms=5000)
        lines = _page_text(page)

        for line in lines:
            if "2027" in line and any(k in line for k in ("校招", "实习", "招聘", "提前批")):
                if line not in seen:
                    seen.add(line)
                    results.append(InternPosition(
                        company="拼多多", title=line, location="上海",
                        link="https://careers.pinduoduo.com/campus",
                        source_note="2027届校招信息",
                    ))

        if not results:
            results.append(InternPosition(
                company="拼多多", title="2027届校招提前批/实习项目已启动",
                location="上海", link="https://careers.pinduoduo.com/campus",
                source_note="具体岗位请在官网查看",
            ))
    except Exception as e:
        results.append(InternPosition(company="拼多多", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


def scrape_bytedance(session: BrowserSession) -> list:
    """字节跳动 - 实习岗位（进入职位列表页 + 滚动加载）"""
    results = []
    seen = set()
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://jobs.bytedance.com/campus/", wait_ms=2500)

        # 进入职位列表页
        navigated = False
        try:
            pos_link = page.query_selector('a[href="/campus/position"]')
            if pos_link:
                with page.expect_navigation(wait_until="domcontentloaded", timeout=15000):
                    pos_link.click()
                navigated = True
        except Exception:
            pass
        if not navigated:
            try:
                page.goto("https://jobs.bytedance.com/campus/position", wait_until="domcontentloaded", timeout=15000)
                navigated = True
            except Exception:
                pass

        page.wait_for_timeout(4000)

        # 点击"实习"筛选，只显示实习岗位
        try:
            el = page.query_selector("text=实习")
            if el:
                el.click()
                page.wait_for_timeout(3000)
        except Exception:
            pass

        _scroll_to_bottom(page, step_ms=250, max_steps=15)
        page.wait_for_timeout(2000)

        lines = _page_text(page)
        skip_words = ["面向全体在校生", "为符合岗位要求的同学提供", "日常实习：", "职位 ID：",
                       "招聘项目", "项目说明", "技术人才项目", "一键投递",
                       "字节范", "字节跳动的使命", "我们因使命",
                       "团队介绍", "筛选", "清除", "职位类别", "工作地点", "更多",
                       "开启新的工作", "搜索", "校招", "首页", "登录"]

        # 岗位行的下一行是职位描述行（含"职位 ID"），据此识别实习岗位
        for i, line in enumerate(lines):
            if i + 1 >= len(lines):
                break
            nxt = lines[i + 1]
            # 描述行特征：包含"职位 ID"和"实习/正式"
            if "职位 ID" not in nxt:
                continue
            if "实习" not in nxt:
                continue  # 只保留实习岗位（排除正式岗）
            if any(s in line for s in skip_words):
                continue
            if len(line) < 6 or len(line) > 100:
                continue

            # 提取地点（从描述行找城市）
            location = _extract_cities(nxt) or "北京/上海/深圳/杭州/广州等"
            # 清理标题
            title = re.sub(r'\s*职位 ID：[A-Z0-9]+', '', line).strip()

            key = title[:20]
            if key not in seen and len(title) > 4:
                seen.add(key)
                results.append(InternPosition(
                    company="字节跳动", title=title, location=location,
                    link="https://jobs.bytedance.com/campus/",
                ))

        # 兜底：从筛选标签行提取招聘项目
        if not results:
            for line in lines:
                if re.match(r'^(前沿技术领域人才|Seed大模型人才).*实习招聘$', line):
                    results.append(InternPosition(
                        company="字节跳动", title=line,
                        location="北京/上海/深圳/杭州/广州等",
                        link="https://jobs.bytedance.com/campus/",
                    ))

        if not results:
            results.append(InternPosition(
                company="字节跳动", title="校园招聘/实习项目已开放",
                location="北京/上海/深圳/杭州/广州等",
                link="https://jobs.bytedance.com/campus/",
                source_note="请在官网注册后查看具体岗位",
            ))
    except Exception as e:
        results.append(InternPosition(company="字节跳动", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载超时"))
    finally:
        page.close()
    return results


def scrape_meituan(session: BrowserSession) -> list:
    """美团 - 校园招聘/实习"""
    results = []
    seen = set()
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://zhaopin.meituan.com/campus", wait_ms=3000)

        # 点击"校园招聘"标签
        try:
            campus_tab = page.query_selector("text=校园招聘")
            if campus_tab:
                campus_tab.click()
                page.wait_for_timeout(4000)
        except Exception:
            pass

        # 点击"实习"筛选
        try:
            intern_filter = page.query_selector("text=实习")
            if intern_filter:
                intern_filter.click()
                page.wait_for_timeout(3000)
        except Exception:
            pass

        lines = _page_text(page)
        skip_words = [
            "首页", "LongCat人才招聘", "北斗计划", "社会招聘", "了解美团",
            "Global Careers", "登录", "筛选", "清除所有", "所在城市", "所在海外国家/地区",
            "搜索更多城市", "职位类别", "部门", "岗位职责", "美团有多少年",
            "更多疑问请查看", "应届校招", "全部校招职位", "日常实习",
            "1.负责", "2.负责", "3.负责", "4.负责",
        ]

        for i, line in enumerate(lines):
            if "实习" not in line:
                continue
            if any(s in line for s in skip_words):
                continue
            if len(line) < 6 or len(line) > 60:
                continue
            if re.match(r'^\d+$', line):
                continue

            location = "北京/上海/深圳/成都等"
            if i + 1 < len(lines):
                cities = _extract_cities(lines[i + 1])
                if cities:
                    location = cities

            key = line[:20]
            if key not in seen:
                seen.add(key)
                results.append(InternPosition(
                    company="美团", title=line, location=location,
                    link="https://zhaopin.meituan.com/campus",
                ))

        if not results:
            results.append(InternPosition(
                company="美团", title="2026春招/实习招聘进行中（转正实习、日常实习）",
                location="北京/上海/深圳/成都等",
                link="https://zhaopin.meituan.com/campus",
                source_note="请在官网筛选查看具体岗位",
            ))
    except Exception as e:
        results.append(InternPosition(company="美团", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


def scrape_alibaba(session: BrowserSession) -> list:
    """阿里巴巴 - 实习信息"""
    results = []
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://talent.alibaba.com/campus/home", wait_ms=5000)
        lines = _page_text(page)

        for line in lines:
            if "实习" in line and len(line) > 2:
                results.append(InternPosition(
                    company="阿里巴巴", title=line,
                    location="杭州/北京/上海/深圳/广州",
                    link="https://talent.alibaba.com/campus/home",
                    source_note="请在官网登录后查看",
                ))

        if not results:
            results.append(InternPosition(
                company="阿里巴巴", title="校园招聘进行中",
                location="杭州/北京/上海/深圳/广州",
                link="https://talent.alibaba.com/campus/home",
                source_note="具体岗位请登录官网查看",
            ))
    except Exception as e:
        results.append(InternPosition(company="阿里巴巴", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


def scrape_baidu(session: BrowserSession) -> list:
    """百度 - 实习信息"""
    results = []
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://talent.baidu.com/campus/", wait_ms=3000)
        lines = _page_text(page)

        for line in lines:
            if "实习" in line and len(line) > 2:
                results.append(InternPosition(
                    company="百度", title=line,
                    location="北京/上海/深圳",
                    link="https://talent.baidu.com/campus/",
                ))

        if not results:
            results.append(InternPosition(
                company="百度", title="校园招聘进行中（需登录查看）",
                location="北京/上海/深圳",
                link="https://talent.baidu.com/campus/",
                source_note="需登录百度账号查看具体岗位",
            ))
    except Exception as e:
        results.append(InternPosition(company="百度", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


def scrape_kuaishou(session: BrowserSession) -> list:
    """快手 - 实习信息"""
    results = []
    page = session.new_page()
    try:
        _goto_with_retry(page, "https://zhaopin.kuaishou.cn/#/campus", wait_ms=5000)
        lines = _page_text(page)

        for line in lines:
            if "实习" in line and len(line) > 2:
                results.append(InternPosition(
                    company="快手", title=line,
                    location="北京/深圳",
                    link="https://zhaopin.kuaishou.cn/#/campus",
                ))

        if not results:
            results.append(InternPosition(
                company="快手", title="校园招聘/日常实习进行中",
                location="北京/深圳",
                link="https://zhaopin.kuaishou.cn/#/campus",
                source_note="具体岗位请在官网查看",
            ))
    except Exception as e:
        results.append(InternPosition(company="快手", title=f"抓取异常: {str(e)[:60]}", source_note="页面加载失败"))
    finally:
        page.close()
    return results


# ─── 爬虫注册表 ─────────────────────────────────────────────

SCRAPERS = [
    ("xiaohongshu", "小红书", scrape_xiaohongshu),
    ("tencent", "腾讯", scrape_tencent),
    ("pinduoduo", "拼多多", scrape_pinduoduo),
    ("bytedance", "字节跳动", scrape_bytedance),
    ("meituan", "美团", scrape_meituan),
    ("alibaba", "阿里巴巴", scrape_alibaba),
    ("baidu", "百度", scrape_baidu),
    ("kuaishou", "快手", scrape_kuaishou),
]


# ─── 报告生成 ───────────────────────────────────────────────

def generate_report(all_positions: list) -> str:
    """生成 Markdown 报告"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_cn = datetime.now().strftime("%Y年%m月%d日")

    lines = ["# 🎓 互联网大厂 2027届实习岗位信息汇总", "",
             f"> **更新日期：{date_cn}** | 自动抓取于 {now}",
             "> 数据来源：各公司官方校园招聘网站",
             "> ⚠️ 部分公司需登录后查看完整岗位，以下信息仅供参考", "",
             "| 公司 | 岗位/项目名称 | 工作地点 | 岗位要求 | 薪资 | 备注 |",
             "|------|-------------|---------|---------|------|------|"]

    for company in COMPANY_ORDER:
        positions = [p for p in all_positions if p.company == company]
        if not positions:
            lines.append(f"| **{company}** | 暂未获取到数据 | - | - | - | 请访问官网查看 |")
            continue
        for idx, pos in enumerate(positions):
            company_display = f"**{company}**" if idx == 0 else ""
            req_short = pos.requirement[:60] + "..." if len(pos.requirement) > 60 else pos.requirement
            note = pos.source_note or ""
            link_display = f"[官网]({pos.link})" if pos.link else ""
            lines.append(f"| {company_display} | {pos.title} | {pos.location} | {req_short} | {pos.salary} | {note} {link_display} |".replace("\n", " "))

    lines += ["", "---", "*报告由自动化脚本自动生成，具体信息以各公司官网为准*"]
    return "\n".join(lines)


def generate_excel(all_positions: list, filepath: str):
    """生成格式化的 Excel 报告"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "实习岗位信息"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    company_font = Font(name="微软雅黑", bold=True, size=10)
    normal_font = Font(name="微软雅黑", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # 标题行
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value=f"互联网大厂 2027届实习岗位信息汇总（{datetime.now().strftime('%Y-%m-%d')}）")
    title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 表头
    headers = ["公司", "岗位/项目名称", "工作地点", "岗位要求（精简）", "薪资范围", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 数据行
    row = 3
    is_alt = False
    for company in COMPANY_ORDER:
        positions = [p for p in all_positions if p.company == company]
        if not positions:
            cells = [company, "暂未获取到数据", "-", "-", "-", "请访问官网查看"]
            for col, val in enumerate(cells, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = company_font if col == 1 else normal_font
                cell.alignment = wrap_align if col in (2, 4, 6) else center_align
                cell.border = thin_border
                if is_alt:
                    cell.fill = alt_fill
            row += 1
            is_alt = not is_alt
            continue

        for idx, pos in enumerate(positions):
            cells = [
                company if idx == 0 else "",
                pos.title,
                pos.location,
                pos.requirement[:80] + "..." if len(pos.requirement) > 80 else pos.requirement,
                pos.salary,
                f"{pos.source_note} {'→' + pos.link if pos.link else ''}",
            ]
            for col, val in enumerate(cells, 1):
                if col == 1 and idx > 0 and len(positions) > 1:
                    continue
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = company_font if col == 1 else normal_font
                cell.alignment = wrap_align if col in (2, 4, 6) else center_align
                cell.border = thin_border
                if is_alt:
                    cell.fill = alt_fill
            if idx == 0 and len(positions) > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row + len(positions) - 1, end_column=1)
            row += 1
        is_alt = not is_alt

    # 列宽
    for i, w in enumerate([14, 42, 24, 50, 10, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 底部说明
    note_cell = ws.cell(row=row + 1, column=1, value="报告由自动化脚本自动生成，具体信息以各公司官网为准")
    note_cell.font = Font(name="微软雅黑", italic=True, size=9, color="999999")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)

    wb.save(filepath)
    print(f"\n📊 Excel 报告已保存至: {filepath}")


# ─── 邮件发送 ───────────────────────────────────────────────

def send_email(report: str, excel_path: Optional[str] = None,
               all_positions: Optional[list] = None, smtp_config: dict = None):
    """通过邮件发送报告（支持 Excel 附件，正文为 HTML 表格）"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication

    cfg = smtp_config or {
        "host": os.environ.get("SMTP_HOST", "smtp.qq.com"),
        "port": int(os.environ.get("SMTP_PORT", 465)),
        "user": os.environ.get("SMTP_USER", ""),
        "password": os.environ.get("SMTP_PASSWORD", ""),
        "to": os.environ.get("SMTP_TO", ""),
    }

    if not cfg["user"] or not cfg["password"]:
        print("[邮件] 未配置 SMTP 账号密码，跳过发送")
        return

    msg = MIMEMultipart()
    msg["Subject"] = f"互联网大厂实习信息日报 - {datetime.now().strftime('%Y-%m-%d')}"
    msg["From"] = cfg["user"]
    msg["To"] = cfg["to"]

    # 正文：真正的 HTML 表格
    if all_positions:
        rows_html = []
        for p in all_positions:
            req = p.requirement[:60] + ("..." if len(p.requirement) > 60 else "")
            rows_html.append(
                f"<tr><td style='padding:6px;border:1px solid #ddd'>{p.company}</td>"
                f"<td style='padding:6px;border:1px solid #ddd'>{p.title}</td>"
                f"<td style='padding:6px;border:1px solid #ddd'>{p.location}</td>"
                f"<td style='padding:6px;border:1px solid #ddd'>{req}</td>"
                f"<td style='padding:6px;border:1px solid #ddd'>{p.salary}</td></tr>"
            )
        html_content = f"""
<html><body style="font-family:微软雅黑,sans-serif;font-size:13px">
<h3>🎓 互联网大厂 2027届实习岗位信息汇总（{datetime.now().strftime('%Y-%m-%d')}）</h3>
<table style="border-collapse:collapse;width:100%">
<tr style="background:#4472C4;color:#fff">
<th style="padding:6px;border:1px solid #ddd">公司</th>
<th style="padding:6px;border:1px solid #ddd">岗位/项目</th>
<th style="padding:6px;border:1px solid #ddd">地点</th>
<th style="padding:6px;border:1px solid #ddd">要求</th>
<th style="padding:6px;border:1px solid #ddd">薪资</th>
</tr>
{''.join(rows_html)}
</table>
<p style="color:#999;font-size:11px">报告由自动化脚本自动生成，具体信息以各公司官网为准。</p>
</body></html>"""
    else:
        html_content = report.replace("\n", "<br>\n")
    msg.attach(MIMEText(html_content, "html", "utf-8"))

    # 附件：Excel
    if excel_path and os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            part = MIMEApplication(f.read())
            part.add_header("Content-Disposition", "attachment",
                            filename=os.path.basename(excel_path))
            msg.attach(part)

    try:
        if cfg["port"] == 465:
            server = smtplib.SMTP_SSL(cfg["host"], cfg["port"])
        else:
            server = smtplib.SMTP(cfg["host"], cfg["port"])
            server.starttls()
        server.login(cfg["user"], cfg["password"])
        server.send_message(msg)
        server.quit()
        print(f"[邮件] 已发送至 {cfg['to']}" + ("（含Excel附件）" if excel_path else ""))
    except Exception as e:
        print(f"[邮件] 发送失败: {e}")


# ─── 主流程 ───────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="互联网大厂实习信息爬虫（优化版）")
    parser.add_argument("--output", "-o", type=str, help="输出文件路径（.md 或 .xlsx）")
    parser.add_argument("--output-excel", type=str, help="单独指定 Excel 输出路径")
    parser.add_argument("--send-email", action="store_true", help="发送邮件（自动附带Excel附件）")
    parser.add_argument("--company", "-c", type=str, help="只抓取指定公司（逗号分隔）")
    args = parser.parse_args()

    print("=" * 60)
    print("  互联网大厂 2027届实习岗位信息爬虫（优化版）")
    print(f"  启动时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    scrapers_to_run = SCRAPERS
    if args.company:
        companies = [c.strip() for c in args.company.split(",")]
        scrapers_to_run = [(k, n, f) for k, n, f in SCRAPERS if n in companies]

    all_positions = []
    start_all = time.time()

    # 复用浏览器会话
    with BrowserSession() as session:
        for key, name, scraper_fn in scrapers_to_run:
            print(f"\n▶ 正在抓取 {name}...", end=" ", flush=True)
            t0 = time.time()
            try:
                positions = scraper_fn(session)
                all_positions.extend(positions)
                print(f"✓ 获取到 {len(positions)} 条信息（{time.time()-t0:.1f}s）")
                for p in positions[:3]:
                    print(f"    - {p.title}")
            except Exception as e:
                print(f"✗ 失败: {e}")
                all_positions.append(InternPosition(company=name, title="抓取失败", source_note=str(e)[:60]))

    print(f"\n{'=' * 60}")
    print(f"  共获取 {len(all_positions)} 条信息，总耗时 {time.time()-start_all:.1f}s")

    # 生成报告
    report = generate_report(all_positions)
    print(f"\n{report}")

    excel_path = None
    # 输出到文件（根据扩展名选择格式）
    if args.output:
        if args.output.endswith(".xlsx"):
            generate_excel(all_positions, args.output)
            excel_path = args.output
        else:
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(report)
            print(f"\n📄 报告已保存至: {args.output}")

    # 单独输出 Excel
    if args.output_excel:
        generate_excel(all_positions, args.output_excel)
        excel_path = args.output_excel

    # 发送邮件（自动附带 Excel 附件）
    if args.send_email:
        send_email(report, excel_path=excel_path, all_positions=all_positions)

    print(f"\n✅ 完成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
